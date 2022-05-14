"""
    Text typesetter for Appareden.
    Truncates ORFIELD text that exceeds limits.
"""
from appareden.rominfo import WAITS, MSGS
from appareden.rominfo import DUMP_XLS_PATH, MAX_LENGTH
from appareden.utils import typeset, ending_typeset, sjis_punctuate, WAITS, NAMES
#from appareden.reinsert import MSGS

from romtools.dump import DumpExcel
from openpyxl.styles import PatternFill

Dump = DumpExcel(DUMP_XLS_PATH)

filenames = ['ORFIELD.EXE', 'ORBTL.EXE']

def safe_print(s, fileObj=None):
    if '\u014d' in s:
        s = s.replace('\u014d', '[o]')
    if '\u016b' in s:
        s = s.replace('\u016b', '[u]')
    print(s)
    if fileObj:
        fileObj.write(s)
        fileObj.write("\n")
        


def starts_with_nametag(s):
    for n in NAMES:
        if s.split('[LN]')[0] == n:
            if s.split('[LN]')[1] != '':
                return True
    return False

# Typesetting EXE strings.
# Currently disabled, since it's eating some strings

"""
for f in filenames:
    rownum = 0
    worksheet = Dump.workbook.get_sheet_by_name(f)
    first_row = list(worksheet.rows)[0]

    header_values = [t.value for t in first_row]
    #jp_col = header_values.index('Japanese')
    en_col = header_values.index('English (Ingame)')
    category_col = header_values.index('Category')

    for row in worksheet.rows:
        if row == first_row:
            continue

        english =row[en_col].value
        display_english = english
        # TDOO: Account for control codes [o], [|], etc. in display_english.
        category = row[category_col].value

        if category:
            maxlen = MAX_LENGTH[category]
            if len(display_english) > maxlen:
                try:
                    print(english, "is too long")
                except UnicodeEncodeError:
                    print("Something with an overline in it is too long")
                while len(english) >= maxlen:
                    english = english[:-1]
                english += "."
                #print(english)
                row[en_col].value = english
                #print()
"""

# Typesetting MSG strings

#msg_files = [f for f in os.listdir(os.path.join('original', 'OR')) if f.endswith('MSG') and not f.startswith('ENDING')]
msgs_to_typeset = MSGS
rownum = 0
worksheet = Dump.workbook['MSG']
first_row = list(worksheet.rows)[0]
header_values = [t.value for t in first_row]
en_col = header_values.index('English (Ingame)')
jp_col = header_values.index('Japanese')
portrait_col = header_values.index('Portrait')
en_typeset_col = header_values.index('English (Typeset)')
file_col = header_values.index('File')

overflows = 0

most_recent_nametag = None

f = open("typeset_script.txt", "w")

for m in msgs_to_typeset:

    #portrait_window_counter = 0
    for row in worksheet.rows:
        nametag = False
        file = row[file_col].value
        terminal_newline = False
        if file == m:
            japanese = row[jp_col].value
            english = row[en_col].value
            portrait = row[portrait_col].value

            if english is None:
                continue

            if (file == "ENDING.MSG"):
                english = ending_typeset(english)
                row[en_typeset_col].value = english
                #input()
                continue

           # safe_print(english, f)  

            # If a character with a portrait is given a nametag in this line,
            # the next line needs to be typeset more aggressively due to less screen space.

            long_names = ['Ultimate Benkei', 'Thunder Dragon', 'Sacrosanct Dragon']

            windows = []
            if '[SPLIT]' in english:
                windows = english.split('[SPLIT]')
            else:
                windows = [english,]

            for i, window in enumerate(windows):

                # If it's not "dialogue", (whispering), or *onomatopaeia*, 
                # it's a nametag
                if window.count('"') == 0 and window.count("(") == 0 and window.count("*") == 0:
                    nametag = True
                    most_recent_nametag = window.rstrip('[LN]')
                    #print("-"*57)

                # For Haley's lines
                window = sjis_punctuate(window)

                #safe_print(window, f)

                #if portrait_window_counter > 0:
                if portrait:
                    window = typeset(window, 36)
                else:
                    window = typeset(window, 57)

                #if portrait in [68000, 77001, ]

                for line in window.split("[LN]"):
                    print(line)
                    #window = window.replace("")
                    if line.replace("[LN]", "") in long_names:
                        name = line.replace("[LN]", "")
                        window = window.replace(name, name + "[LN][LN]")
                        #print(line)
                        #input()


                #safe_print(window, f)

                window_lines = window.split('[LN]')

                #safe_print(",".join(window_lines), f)

                # Terminal [LN] sometimes messes it up, so ignore those
                if window_lines[-1] == '':
                    window_lines = window_lines[:-1]
                    terminal_newline = True
                #print("Removed an extraneous line in row %s" % window)

                line_count = 5

                if i > 0:
                    safe_print("%s%s" % (" "*20, most_recent_nametag), f)

                if starts_with_nametag(window):
                    #print("Starts with nametag")
                    window_lines[1] = window_lines[1].lstrip()

                for e in window_lines:

                    # Remove WAIT control codes when printing here
                    for w in WAITS:
                        e = e.replace(w, '')

                    if portrait:
                        safe_print("%s%s" % (" "*20, e), f)
                    else:
                        safe_print(e, f)
                    line_count -= 1

                if not nametag:

                    while line_count > 0:
                        safe_print("", f)
                        line_count -= 1

                    safe_print('-'*57, f)

                    # Prevent windows with a nametag prefix from giving a false-positive "overflow"
                    if starts_with_nametag(window):
                        #safe_print("Starts with nametag", f)
                        #window_lines[1] = window_lines[1].lstrip()
                        line_count += 1

                    if line_count < 0:
                        # Mark the overflowing cells with a mint green background
                        safe_print("^ This window overflows\n", f)
                        row[en_col].fill = PatternFill(fgColor="c6ffe2", fill_type='solid')
                        overflows += 1
                    else:
                        # Need to re-whiten cells that aren't overflowing.
                        # This removes the marking from ones that were fixed
                        row[en_col].fill = PatternFill(fgColor='ffffff', fill_type='solid')

                # Save the changes to windows
                windows[i] = '[LN]'.join(window_lines)

            #if terminal_newline:
            #    english += '[LN]'

            english = '[SPLIT]'.join(windows)

            # Add a terminal newline if it's missing in English
            if english is not None and japanese is not None:
                if japanese.endswith('[LN]') and not english.endswith('[LN]'):
                    english += '[LN]'
                #print("There should be a terminal [LN] here")

            #if 'Ultimate Benkei' in english:
            #    print(english)
            #    input()

            row[en_typeset_col].value = english

Dump.workbook.save(DUMP_XLS_PATH)

safe_print("%s windows overflow" % overflows, f)